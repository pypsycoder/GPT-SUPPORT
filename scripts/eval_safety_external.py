"""
Замер L0 на внешнем датасете — данные, которых детектор точно не видел.

Датасет: «Dataset of depressive and suicidal posts», Narynov, Mukhtarkhanuly,
Omarov (2020), Mendeley Data 10.17632/838dbcjpxb.1, CC BY 4.0.
64 040 постов ВКонтакте, размечены психиатрами республиканского
научно-практического центра психического здоровья. Файл в LLM_test/external/
(каталог gitignored, в репозиторий данные не попадают).

ЗАЧЕМ ИМЕННО ОН. Ночной синтетический набор писала та же модель, что стоит
в проде, поэтому цифры на нём меряют её представление о кризисе, а не речь
живых людей. Здесь текст чужой и детектору незнаком.

ЧТО ЗДЕСЬ НЕЛЬЗЯ ИЗМЕРИТЬ. Метка — «депрессивный / не депрессивный», а не
«кризис / не кризис». Депрессивный пост в нашей таксономии чаще всего concern,
а не urgent, поэтому recall по urgent на label=1 мерить бессмысленно: он
покажет не качество детектора, а разницу определений.

ЧТО ИЗМЕРИТЬ МОЖНО:
  1. Ложные срабатывания на label=0 — 32 тысячи настоящих нейтральных текстов.
     Метрика чистая и однозначная: urgent здесь не должен срабатывать почти никогда.
  2. Чувствительность на label=1 — доля постов, где детектор заметил хоть что-то
     (urgent или concern). Это не recall кризиса, а «не прошёл ли мимо целиком».
  3. То же самое для старого keyword-роутера, для сравнения.

Вызовов модели не требует.

Запуск:
    python scripts/eval_safety_external.py [--limit 0] [--show 15]
"""

from __future__ import annotations

import argparse
import io
import random
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.core.config import load_environment

load_environment()

from app.llm import router_l0  # noqa: E402
from app.llm.router import RequestType, classify_request  # noqa: E402

DEFAULT_PATH = ROOT_DIR / "LLM_test" / "external" / "depressive_data.xlsx"

# Посты в соцсети длинные, реплики в чате короткие. Обрезаем, чтобы мерить
# на сопоставимом объёме текста, а не на простыне в тысячу знаков.
MAX_CHARS = 400


def load_rows(path: Path, limit: int = 0) -> list[tuple[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[tuple[str, int]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index == 0:
            continue  # заголовок
        text, label = row[0], row[1]
        if text is None or label is None:
            continue
        text = " ".join(str(text).split())[:MAX_CHARS]
        if not text:
            continue
        try:
            rows.append((text, int(label)))
        except (TypeError, ValueError):
            continue
        if limit and len(rows) >= limit:
            break
    workbook.close()
    return rows


def main() -> None:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", type=Path, default=DEFAULT_PATH)
    parser.add_argument("--limit", type=int, default=0, help="0 = весь файл")
    parser.add_argument("--show", type=int, default=15, help="сколько примеров печатать")
    args = parser.parse_args()

    print(f"читаю {args.path.name} ...")
    rows = load_rows(args.path, args.limit)
    labels = Counter(label for _, label in rows)
    print(f"строк: {len(rows)} | по меткам: {dict(sorted(labels.items()))}")
    print(f"обрезка текста: {MAX_CHARS} символов\n")

    stats: dict[str, Counter] = {"l0": Counter(), "old": Counter()}
    false_alarms: list[tuple[str, str]] = []

    for text, label in rows:
        decision = router_l0.classify(text)
        old_safety = classify_request(text, "text").request_type is RequestType.SAFETY

        stats["l0"][(label, decision.safety_level)] += 1
        stats["old"][(label, "urgent" if old_safety else "none")] += 1

        if label == 0 and decision.safety_level == "urgent":
            false_alarms.append((text, decision.rule or "?"))

    neutral = labels.get(0, 0)
    depressive = labels.get(1, 0)

    print("=== 1. Ложные срабатывания на нейтральных текстах (label=0) ===")
    for name, key in (("L0", "l0"), ("старый роутер", "old")):
        fired = stats[key][(0, "urgent")]
        rate = fired / neutral if neutral else 0.0
        print(f"  {name:16s} urgent: {fired:5d} из {neutral} ({rate:.2%})")
    l0_concern_neutral = stats["l0"][(0, "concern")]
    print(f"  L0 concern на нейтральных: {l0_concern_neutral} ({l0_concern_neutral / neutral:.2%})")

    print("\n=== 2. Чувствительность на депрессивных текстах (label=1) ===")
    print("    (не recall кризиса: метка датасета шире нашего urgent)")
    l0_urgent = stats["l0"][(1, "urgent")]
    l0_concern = stats["l0"][(1, "concern")]
    old_urgent = stats["old"][(1, "urgent")]
    print(f"  L0 urgent:            {l0_urgent:5d} ({l0_urgent / depressive:.2%})")
    print(f"  L0 concern:           {l0_concern:5d} ({l0_concern / depressive:.2%})")
    print(f"  L0 заметил хоть что-то: {l0_urgent + l0_concern:5d} "
          f"({(l0_urgent + l0_concern) / depressive:.2%})")
    print(f"  старый роутер urgent: {old_urgent:5d} ({old_urgent / depressive:.2%})")

    if false_alarms and args.show:
        print(f"\n=== 3. Ложные срабатывания L0 — что чинить ({len(false_alarms)} всего) ===")
        by_rule = Counter(rule for _, rule in false_alarms)
        print(f"  по правилам: {dict(by_rule.most_common())}\n")
        random.seed(0)
        for text, rule in random.sample(false_alarms, min(args.show, len(false_alarms))):
            print(f"  [{rule}] {text[:150]}")


if __name__ == "__main__":
    main()
