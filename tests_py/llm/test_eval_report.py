from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

import pytest

from scripts.run_llm_eval import EvalCase, _build_markdown_report, _filter_cases, _write_workbook


def _sample_result() -> dict:
    return {
        "case_id": "case_1",
        "category": "mixed",
        "status": "PASS",
        "issues": [],
        "input_text": "text",
        "expected_policy": None,
        "notes": "notes",
        "model_tier": "lite",
        "response": "response",
        "tokens_input": 10,
        "tokens_output": 5,
        "response_time_ms": 100,
        "diagnostics": {
            "prompt": {},
            "classify": {
                "request_type": "simple",
                "router_domain": "sleep",
                "effective_domain": "sleep",
            },
            "parser": {
                "mood": None,
                "domain_hints": [],
            },
            "patient_context": {
                "rag": {"backend": "pgvector", "hit_count": 1},
                "rag_context": ["RAG text 1"],
            },
            "llm_call": {"latency_ms": 100},
            "summary": {"total_stage_latency_ms": 150, "fallback_points": []},
            "stages": [
                {"name": "classification", "status": "ok", "latency_ms": 2},
                {"name": "supervisor", "status": "ok", "latency_ms": 90},
            ],
            "supervisor": {
                "execution_kind": "delegate",
                "selected_agents": ["emotion"],
                "needs_clarification": False,
                "graph_path": ["intake_analyze", "delegation_analyze", "finalize_reply"],
            },
        },
    }


def test_write_workbook_includes_current_pipeline_sheets(tmp_path: Path):
    output_path = tmp_path / "eval.xlsx"

    _write_workbook([_sample_result()], output_path, patient_id=1)

    wb = load_workbook(output_path)
    assert "Cases" in wb.sheetnames
    assert "Timeline" in wb.sheetnames
    assert "Diagnostics" in wb.sheetnames
    assert "Orchestration" not in wb.sheetnames

    cases_headers = [cell.value for cell in wb["Cases"][1]]
    assert "supervisor_execution" in cases_headers
    assert "supervisor_selected_agents" in cases_headers

    timeline_ws = wb["Timeline"]
    timeline_headers = [cell.value for cell in timeline_ws[1]]
    assert "step_no" in timeline_headers
    assert "step_label" in timeline_headers
    assert "details_json" in timeline_headers
    timeline_labels = [cell.value for cell in timeline_ws["E"]]
    assert "classification" in timeline_labels
    assert "supervisor_graph" in timeline_labels


def test_filter_cases_returns_single_requested_case():
    cases = [
        EvalCase("case_a", "general", "text a", None, [], [], ""),
        EvalCase("case_b", "general", "text b", None, [], [], ""),
    ]

    filtered = _filter_cases(cases, "case_b")

    assert [item.case_id for item in filtered] == ["case_b"]


def test_filter_cases_raises_for_unknown_case():
    cases = [EvalCase("case_a", "general", "text a", None, [], [], "")]

    with pytest.raises(RuntimeError, match="Case 'missing_case' not found"):
        _filter_cases(cases, "missing_case")


def test_build_markdown_report_includes_current_pipeline_timeline():
    report = _build_markdown_report([_sample_result()], generated_at="2026-04-05 11:30:00", patient_id=1)

    assert "#### case_1" in report
    assert "**RAG Context**" in report
    assert "1. RAG text 1" in report
    assert "**Timeline**" in report
    assert "`classification` / `pipeline`" in report
    assert "`supervisor_graph` / `supervisor`" in report
