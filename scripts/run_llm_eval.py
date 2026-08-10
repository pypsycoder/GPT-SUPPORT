from __future__ import annotations

import argparse
import asyncio
from collections import Counter
from dataclasses import dataclass, replace
from datetime import datetime
import json
from pathlib import Path
import re
from statistics import mean
import sys

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.core.config import load_environment


DEFAULT_CASES_PATH = ROOT_DIR / "LLM_test" / "cases" / "chat_quality_cases.yaml"
DEFAULT_REPORTS_DIR = ROOT_DIR / "LLM_test" / "reports"

HYDRATION_PHRASES = (
    "пить больше воды",
    "пить воду",
    "пить больше жидкости",
    "пить жидкость",
    "восполнять жидкость",
    "hydrated",
)

HYDRATION_WORD_STEMS = (
    "гидрат",
)

FOOD_PHRASES = (
    "что съесть",
    "что выпить",
    "тяжелой еды",
    "легкую пищу",
    "легкая пища",
    "напитков с кофеином",
    "напитки с кофеином",
    "травяной чай",
)

FOOD_WORDS = (
    "перекус",
    "перекуси",
    "перекусить",
    "поесть",
    "поешь",
    "кушать",
    "кушай",
    "кофеин",
    "кофе",
    "чай",
    "напиток",
    "напитки",
    "суп",
    "йогурт",
    "кефир",
    "фрукты",
    "еда",
    "пища",
)

FOOD_WORD_STEMS = (
    "ромашк",
    "мятн",
)
TEMPLATE_REASSURANCE_PATTERNS = (
    "ты справишься",
    "держись",
    "всё будет хорошо",
    "все будет хорошо",
    "не переживай",
)
DOCTOR_FIRST_PATTERNS = (
    "обратись к врачу",
    "обратитесь к врачу",
    "обратись к медсестре",
    "расскажи врачу",
    "расскажи медсестре",
    "поговори с врачом",
    "поговори с медсестрой",
    "сообщи врачу",
    "сообщи медсестре",
    "сообщи медицинскому персоналу",
)
CARE_TEAM_ASSUMPTION_PATTERNS = (
    "медсестра поможет",
    "врач поможет",
    "они помогут",
    "точно помогут",
    "они подскажут",
    "подскажут, что делать",
    "они знают, как помочь",
    "они умеют помогать",
    "они знают, что делать",
    "тебе помогут",
)
ACTION_PATTERNS = (
    "попробуй",
    "сделай",
    "отдохни",
    "снизь",
    "замедлись",
    "подыши",
    "закрой глаза",
    "сделай паузу",
    "отложи",
    "избегай",
    "постарайся",
    "выбери",
    "найди",
    "используй",
    "ограни",
    "запиши",
    "сообщи",
    "обсуди",
    "скажи",
    "поговори",
    "обратись",
    "поешь",
    "поесть",
    "выпей",
    "выпить",
)

ACTION_WORD_STEMS = (
    "созда",
    "убер",
    "постара",
)

STATUS_FILL = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "WARN": PatternFill("solid", fgColor="FFEB9C"),
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
}


@dataclass
class EvalCase:
    case_id: str
    category: str
    text: str
    expected_policy: str | None
    forbid_checks: list[str]
    require_checks: list[str]
    notes: str


def _load_cases(path: Path) -> list[EvalCase]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    cases: list[EvalCase] = []
    for item in payload.get("cases", []):
        checks = item.get("checks", {})
        cases.append(
            EvalCase(
                case_id=item["id"],
                category=item.get("category", "general"),
                text=item["text"],
                expected_policy=item.get("expected_policy"),
                forbid_checks=list(checks.get("forbid", [])),
                require_checks=list(checks.get("require", [])),
                notes=item.get("notes", ""),
            )
        )
    return cases


def _filter_cases(cases: list[EvalCase], case_id: str | None) -> list[EvalCase]:
    if not case_id:
        return cases
    filtered = [case for case in cases if case.case_id == case_id]
    if not filtered:
        raise RuntimeError(f"Case '{case_id}' not found in loaded eval cases.")
    return filtered


def _contains_any(text: str, patterns: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(pattern in lowered for pattern in patterns)


def _contains_word(text: str, words: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(re.search(rf"\b{re.escape(word)}\b", lowered) for word in words)


def _contains_word_stem(text: str, stems: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(re.search(rf"\b{re.escape(stem)}\w*", lowered) for stem in stems)


def _contains_hydration_advice(text: str) -> bool:
    return _contains_any(text, HYDRATION_PHRASES) or _contains_word_stem(text, HYDRATION_WORD_STEMS)


def _contains_food_advice(text: str) -> bool:
    return (
        _contains_any(text, FOOD_PHRASES)
        or _contains_word(text, FOOD_WORDS)
        or _contains_word_stem(text, FOOD_WORD_STEMS)
    )


def _starts_with_any_line(text: str, patterns: tuple[str, ...]) -> bool:
    lines = [line.strip(" -\t").lower() for line in text.splitlines() if line.strip()]
    if not lines:
        return False
    first_line = lines[0]
    return any(pattern in first_line for pattern in patterns)


def _has_early_escalation(text: str, patterns: tuple[str, ...], *, window: int = 3) -> bool:
    lines = [line.strip(" -\t").lower() for line in text.splitlines() if line.strip()]
    for line in lines[:window]:
        if any(pattern in line for pattern in patterns):
            return True
    return False


def _detect_issues(response_text: str, diagnostics: dict, case: EvalCase) -> list[str]:
    issues: list[str] = []
    prompt_diag = diagnostics.get("prompt", {})

    if case.expected_policy and prompt_diag.get("selected_policy") and prompt_diag.get("selected_policy") != case.expected_policy:
        issues.append("wrong_policy")

    checks = {
        "hydration_advice": _contains_hydration_advice(response_text),
        "food_advice": _contains_food_advice(response_text),
        "template_reassurance": _contains_any(response_text, TEMPLATE_REASSURANCE_PATTERNS),
        "doctor_first": _starts_with_any_line(response_text, DOCTOR_FIRST_PATTERNS),
        "early_escalation": _has_early_escalation(response_text, DOCTOR_FIRST_PATTERNS),
        "care_team_assumption": _contains_any(response_text, CARE_TEAM_ASSUMPTION_PATTERNS),
        "action_step": _contains_any(response_text, ACTION_PATTERNS) or _contains_word_stem(response_text, ACTION_WORD_STEMS),
    }

    if checks["care_team_assumption"]:
        issues.append("care_team_assumption")

    for forbid in case.forbid_checks:
        if checks.get(forbid, False):
            issues.append(forbid)

    for required in case.require_checks:
        if required == "action_step" and not checks["action_step"]:
            issues.append("no_action_step")

    return issues


def _status_from_issues(issues: list[str]) -> str:
    if not issues:
        return "PASS"
    if any(issue != "no_action_step" for issue in issues):
        return "FAIL"
    return "WARN"


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 60)


def _safe_mean(values: list[int]) -> float:
    return float(mean(values)) if values else 0.0


def _json_dump(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


_SUPERVISOR_STEPS = ("intake", "delegation", "expert")


def _supervisor_llm_steps(supervisor: dict) -> list[dict]:
    """Диагностика LLM по каждому узлу графа (intake / delegation / expert)."""
    steps = []
    for name in _SUPERVISOR_STEPS:
        llm = ((supervisor.get(name) or {}).get("llm") or {})
        if llm:
            steps.append(llm)
    return steps


def _supervisor_parse_mode(supervisor: dict) -> str:
    """field_block (легаси) | structured (response_format json_schema)."""
    modes = {str(llm.get("parse_mode") or "field_block") for llm in _supervisor_llm_steps(supervisor)}
    if not modes:
        return "-"
    return ", ".join(sorted(modes))


def _supervisor_parse_metric(supervisor: dict, key: str) -> int:
    """Сумма счётчика по узлам хода: попытки парса и починки схемы."""
    return sum(int(llm.get(key) or 0) for llm in _supervisor_llm_steps(supervisor))


def _normalize_runtime_diagnostics(diagnostics: dict, *, router_result, response) -> dict:
    normalized = dict(diagnostics or {})
    normalized.setdefault(
        "classify",
        {
            "request_type": router_result.request_type.value,
            "router_domain": router_result.domain_hint,
            "effective_domain": router_result.domain_hint,
        },
    )
    normalized.setdefault("parser", {"mood": None, "domain_hints": []})
    normalized.setdefault("prompt", {})
    normalized.setdefault("patient_context", {"rag": {}, "rag_context": []})
    normalized.setdefault(
        "llm_call",
        {
            "latency_ms": response.response_time_ms,
            "tokens_input": response.tokens_input,
            "tokens_output": response.tokens_output,
            "account_id": response.account_id,
        },
    )
    normalized.setdefault(
        "summary",
        {
            "total_stage_latency_ms": normalized.get("total_latency_ms", response.response_time_ms),
            "fallback_points": [],
        },
    )
    return normalized


def _extract_case_timeline_rows(item: dict) -> list[dict[str, object]]:
    diagnostics = item["diagnostics"]
    rows: list[dict[str, object]] = []
    step_no = 1

    def add_row(
        *,
        step_label: str,
        actor: str,
        status: str,
        latency_ms: int = 0,
        tokens_input: int = 0,
        tokens_output: int = 0,
        prompt_chars: int = 0,
        reasons: list[str] | None = None,
        warnings: list[str] | None = None,
        selected_context_sections: list[str] | None = None,
        details: object | None = None,
    ) -> None:
        nonlocal step_no
        rows.append(
            {
                "case_id": item["case_id"],
                "status": item["status"],
                "model_tier": item.get("model_tier", "-"),
                "step_no": step_no,
                "step_label": step_label,
                "actor": actor,
                "step_status": status,
                "latency_ms": latency_ms,
                "tokens_input": tokens_input,
                "tokens_output": tokens_output,
                "prompt_chars": prompt_chars,
                "reasons": ", ".join(reasons or []),
                "warnings": ", ".join(warnings or []),
                "selected_context_sections": ", ".join(selected_context_sections or []),
                "details_json": _json_dump(details or {}),
            }
        )
        step_no += 1

    for stage in diagnostics.get("stages", []):
        add_row(
            step_label=str(stage.get("name") or "stage"),
            actor="pipeline",
            status=str(stage.get("status") or "unknown"),
            latency_ms=int(stage.get("latency_ms", 0) or 0),
            warnings=[str(stage.get("error"))] if stage.get("error") else [],
            details=stage,
        )

    supervisor = diagnostics.get("supervisor", {})
    if supervisor:
        add_row(
            step_label="supervisor_graph",
            actor="supervisor",
            status=str(supervisor.get("execution_kind") or "ok"),
            latency_ms=int(supervisor.get("latency_ms", 0) or 0),
            reasons=list(supervisor.get("selected_agents") or []),
            details={
                "graph_path": supervisor.get("graph_path") or [],
                "needs_clarification": supervisor.get("needs_clarification"),
                "state_delta": supervisor.get("state_delta") or {},
                "llm_totals": supervisor.get("llm_totals") or {},
                "education_grounding": supervisor.get("education_grounding") or {},
            },
        )

    memory = diagnostics.get("memory", {})
    if memory:
        add_row(
            step_label="memory",
            actor="memory_write",
            status=str(memory.get("status") or "ok"),
            details=memory,
        )

    return rows


async def _run_eval_cases(cases: list[EvalCase], patient_id: int, source: str) -> list[dict]:
    from app.llm.pipeline import LLMPipeline, LLMRequest
    from app.llm.router import ModelTier, classify_request
    from core.db.session import async_session_factory

    override_model_tier: str | None = getattr(_run_eval_cases, "_override_model_tier", None)
    pipeline = LLMPipeline()
    results: list[dict] = []
    for case in cases:
        router_result = classify_request(case.text, source)
        if override_model_tier:
            router_result = replace(router_result, model_tier=ModelTier(override_model_tier))
        async with async_session_factory() as session:
            result = await pipeline.process(
                LLMRequest(
                    patient_id=patient_id,
                    user_input=case.text,
                    source=source,
                    router_result=router_result,
                    strict_model_tier=bool(override_model_tier),
                    db=session,
                )
            )
            await session.rollback()

        diagnostics = _normalize_runtime_diagnostics(result.diagnostics, router_result=router_result, response=result)
        issues = _detect_issues(result.response, diagnostics, case)
        results.append(
            {
                "case_id": case.case_id,
                "category": case.category,
                "input_text": case.text,
                "expected_policy": case.expected_policy,
                "notes": case.notes,
                "model_tier": router_result.model_tier.value,
                "status": _status_from_issues(issues),
                "issues": issues,
                "response": result.response,
                "tokens_input": result.tokens_input,
                "tokens_output": result.tokens_output,
                "response_time_ms": result.response_time_ms,
                "diagnostics": diagnostics,
            }
        )
    return results


def _parse_summary_rows(results: list[dict]) -> list[tuple[str, object]]:
    """Метрики шага 3: доля успешных парсов карточек и доля repair-ретраев.

    Один «парс» — один вызов узла графа. Успешным считается узел, который отдал
    карточку (``final_status == success``); ``repair_attempts`` — починки схемы
    внутри ``GigaChatClient.structured()``, целевая доля < 2%.
    """
    steps = [
        llm
        for item in results
        for llm in _supervisor_llm_steps(item["diagnostics"].get("supervisor") or {})
    ]
    if not steps:
        return [("Card parse steps", 0)]

    attempts = sum(int(llm.get("attempts_total") or 0) for llm in steps)
    repairs = sum(int(llm.get("repair_attempts") or 0) for llm in steps)
    succeeded = sum(1 for llm in steps if llm.get("final_status") == "success")
    modes = sorted({str(llm.get("parse_mode") or "field_block") for llm in steps})

    return [
        ("Card parse mode", ", ".join(modes)),
        ("Card parse steps", len(steps)),
        ("Card parse success rate", round(succeeded / len(steps), 4)),
        ("Card parse attempts (incl. retries)", attempts),
        ("Card repair attempts", repairs),
        ("Card repair rate", round(repairs / attempts, 4) if attempts else 0.0),
    ]


def _build_summary_sheet(ws, results: list[dict], generated_at: str, patient_id: int, *, title: str = "Summary") -> None:
    statuses = Counter(item["status"] for item in results)
    issues = Counter(issue for item in results for issue in item["issues"])
    policies = Counter(item["diagnostics"]["prompt"].get("selected_policy", "unknown") for item in results)

    ws.title = title
    ws.append(["Metric", "Value"])
    rows = [
        ("Generated at", generated_at),
        ("Patient ID", patient_id),
        ("Cases total", len(results)),
        ("PASS", statuses.get("PASS", 0)),
        ("WARN", statuses.get("WARN", 0)),
        ("FAIL", statuses.get("FAIL", 0)),
        (
            "Avg total_stage_latency_ms",
            round(_safe_mean([int(item["diagnostics"]["summary"].get("total_stage_latency_ms", 0)) for item in results]), 1),
        ),
        (
            "Avg llm_latency_ms",
            round(_safe_mean([int(item["diagnostics"]["llm_call"].get("latency_ms", 0)) for item in results]), 1),
        ),
        ("Avg tokens_input", round(_safe_mean([int(item["tokens_input"]) for item in results]), 1)),
        ("Avg tokens_output", round(_safe_mean([int(item["tokens_output"]) for item in results]), 1)),
    ]
    rows.extend(_parse_summary_rows(results))
    for row in rows:
        ws.append(list(row))

    issue_row = len(rows) + 3
    ws.cell(issue_row, 1, "Issue")
    ws.cell(issue_row, 2, "Count")
    for offset, (issue, count) in enumerate(sorted(issues.items()), start=1):
        ws.cell(issue_row + offset, 1, issue)
        ws.cell(issue_row + offset, 2, count)

    policy_col = 4
    ws.cell(issue_row, policy_col, "Policy")
    ws.cell(issue_row, policy_col + 1, "Count")
    for offset, (policy, count) in enumerate(sorted(policies.items()), start=1):
        ws.cell(issue_row + offset, policy_col, policy)
        ws.cell(issue_row + offset, policy_col + 1, count)

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _build_cases_sheet(ws, results: list[dict], *, title: str = "Cases") -> None:
    ws.title = title
    headers = [
        "case_id",
        "category",
        "status",
        "issues",
        "input_text",
        "expected_policy",
        "selected_policy",
        "request_type",
        "router_domain",
        "effective_domain",
        "parser_mood",
        "parser_domain_hints",
        "supervisor_execution",
        "supervisor_selected_agents",
        "supervisor_needs_clarification",
        "graph_path",
        "parse_mode",
        "parse_attempts",
        "repair_attempts",
        "response",
        "tokens_input",
        "tokens_output",
        "rag_backend",
        "rag_hit_count",
        "embedding_ms",
        "search_ms",
        "llm_latency_ms",
        "total_stage_latency_ms",
        "notes",
        "model_tier",
    ]
    ws.append(headers)
    for item in results:
        diagnostics = item["diagnostics"]
        supervisor = diagnostics.get("supervisor", {})
        ws.append(
            [
                item["case_id"],
                item["category"],
                item["status"],
                ", ".join(item["issues"]) or "-",
                item["input_text"],
                item["expected_policy"] or "-",
                diagnostics["prompt"].get("selected_policy"),
                diagnostics["classify"].get("request_type"),
                diagnostics["classify"].get("router_domain"),
                diagnostics["classify"].get("effective_domain"),
                diagnostics["parser"].get("mood"),
                ", ".join(diagnostics["parser"].get("domain_hints") or []),
                supervisor.get("execution_kind", "-"),
                ", ".join(supervisor.get("selected_agents") or []),
                supervisor.get("needs_clarification", False),
                " > ".join(supervisor.get("graph_path") or []),
                _supervisor_parse_mode(supervisor),
                _supervisor_parse_metric(supervisor, "attempts_total"),
                _supervisor_parse_metric(supervisor, "repair_attempts"),
                item["response"],
                item["tokens_input"],
                item["tokens_output"],
                diagnostics["patient_context"]["rag"].get("backend"),
                diagnostics["patient_context"]["rag"].get("hit_count"),
                diagnostics["patient_context"]["rag"].get("embedding_request_ms", 0),
                diagnostics["patient_context"]["rag"].get("vector_search_ms", 0),
                diagnostics["llm_call"].get("latency_ms", 0),
                diagnostics["summary"].get("total_stage_latency_ms", 0),
                item["notes"],
                item.get("model_tier", "-"),
            ]
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        fill = STATUS_FILL.get(row[2].value)
        if fill:
            for cell in row:
                cell.fill = fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autosize_columns(ws)

def _build_timeline_sheet(ws, results: list[dict], *, title: str = "Timeline") -> None:
    ws.title = title
    headers = [
        "case_id",
        "status",
        "model_tier",
        "step_no",
        "step_label",
        "actor",
        "step_status",
        "latency_ms",
        "tokens_input",
        "tokens_output",
        "prompt_chars",
        "reasons",
        "warnings",
        "selected_context_sections",
        "details_json",
    ]
    ws.append(headers)
    for item in results:
        for row in _extract_case_timeline_rows(item):
            ws.append([row.get(header, "") for header in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        fill = STATUS_FILL.get(row[1].value)
        if fill:
            for cell in row:
                cell.fill = fill
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _build_diagnostics_sheet(ws, results: list[dict], *, title: str = "Diagnostics") -> None:
    ws.title = title
    ws.append(
        [
            "case_id",
            "model_tier",
            "policy_reasons",
            "fallback_points",
            "available_sections",
            "included_sections",
            "supervisor_execution",
            "supervisor_selected_agents",
            "supervisor_needs_clarification",
            "graph_path",
            "diagnostics_json",
        ]
    )
    for item in results:
        diagnostics = item["diagnostics"]
        supervisor = diagnostics.get("supervisor", {})
        ws.append(
            [
                item["case_id"],
                item.get("model_tier", "-"),
                ", ".join(diagnostics["prompt"].get("policy_reasons") or []),
                ", ".join(diagnostics["summary"].get("fallback_points") or []),
                ", ".join(diagnostics["prompt"].get("available_sections") or []),
                ", ".join(diagnostics["prompt"].get("included_sections") or []),
                supervisor.get("execution_kind", "-"),
                ", ".join(supervisor.get("selected_agents") or []),
                supervisor.get("needs_clarification", False),
                " > ".join(supervisor.get("graph_path") or []),
                _json_dump(diagnostics),
            ]
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _build_compare_sheet(ws, runs: dict[str, list[dict]], generated_at: str, patient_id: int) -> None:
    ws.title = "Compare"
    ws.append(["Metric", "Value"])
    ws.append(["Generated at", generated_at])
    ws.append(["Patient ID", patient_id])
    ws.append(["Tiers compared", ", ".join(runs.keys())])
    ws.append([])
    ws.append(["Tier", "PASS", "WARN", "FAIL", "Avg total_stage_latency_ms", "Avg llm_latency_ms", "Avg tokens_input", "Avg tokens_output"])

    for tier_name, results in runs.items():
        statuses = Counter(item["status"] for item in results)
        ws.append(
            [
                tier_name,
                statuses.get("PASS", 0),
                statuses.get("WARN", 0),
                statuses.get("FAIL", 0),
                round(_safe_mean([int(item["diagnostics"]["summary"].get("total_stage_latency_ms", 0)) for item in results]), 1),
                round(_safe_mean([int(item["diagnostics"]["llm_call"].get("latency_ms", 0)) for item in results]), 1),
                round(_safe_mean([int(item["tokens_input"]) for item in results]), 1),
                round(_safe_mean([int(item["tokens_output"]) for item in results]), 1),
            ]
        )

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for cell in ws[6]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A6"
    _autosize_columns(ws)


def _write_workbook(results: list[dict] | dict[str, list[dict]], output_path: Path, patient_id: int) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(results, dict):
        compare_ws = wb.active
        _build_compare_sheet(compare_ws, results, generated_at, patient_id)
        for tier_name, tier_results in results.items():
            suffix = tier_name.upper()
            _build_summary_sheet(wb.create_sheet(), tier_results, generated_at, patient_id, title=f"Summary_{suffix}")
            _build_cases_sheet(wb.create_sheet(), tier_results, title=f"Cases_{suffix}")
            _build_timeline_sheet(wb.create_sheet(), tier_results, title=f"Timeline_{suffix}")
            _build_diagnostics_sheet(wb.create_sheet(), tier_results, title=f"Diag_{suffix}")
    else:
        summary_ws = wb.active
        _build_summary_sheet(summary_ws, results, generated_at, patient_id)
        _build_cases_sheet(wb.create_sheet(), results)
        _build_timeline_sheet(wb.create_sheet(), results)
        _build_diagnostics_sheet(wb.create_sheet(), results)
    wb.save(output_path)


def _build_markdown_report(results: list[dict] | dict[str, list[dict]], *, generated_at: str, patient_id: int) -> str:
    lines: list[str] = [
        "# LLM Eval Report",
        "",
        f"- Generated at: {generated_at}",
        f"- Patient ID: {patient_id}",
    ]

    def append_run(run_name: str | None, run_results: list[dict]) -> None:
        statuses = Counter(item["status"] for item in run_results)
        issues = Counter(issue for item in run_results for issue in item["issues"])
        if run_name:
            lines.extend(["", f"## Run: {run_name}", ""])
        else:
            lines.append("")
        lines.extend(
            [
                "### Summary",
                "",
                f"- PASS: {statuses.get('PASS', 0)}",
                f"- WARN: {statuses.get('WARN', 0)}",
                f"- FAIL: {statuses.get('FAIL', 0)}",
                f"- Avg total_stage_latency_ms: {round(_safe_mean([int(item['diagnostics']['summary'].get('total_stage_latency_ms', 0)) for item in run_results]), 1)}",
                f"- Avg llm_latency_ms: {round(_safe_mean([int(item['diagnostics']['llm_call'].get('latency_ms', 0)) for item in run_results]), 1)}",
                f"- Avg tokens_input: {round(_safe_mean([int(item['tokens_input']) for item in run_results]), 1)}",
                f"- Avg tokens_output: {round(_safe_mean([int(item['tokens_output']) for item in run_results]), 1)}",
                f"- Issues: {dict(issues) if issues else '{}'}",
                "",
                "### Cases",
                "",
            ]
        )
        for item in run_results:
            diagnostics = item["diagnostics"]
            supervisor = diagnostics.get("supervisor", {})
            rag_context_items = list(diagnostics.get("patient_context", {}).get("rag_context") or [])
            rag_backend = diagnostics.get("patient_context", {}).get("rag", {}).get("backend")
            lines.extend(
                [
                    f"#### {item['case_id']}",
                    "",
                    f"- Status: {item['status']}",
                    f"- Category: {item['category']}",
                    f"- Model tier: {item.get('model_tier', '-')}",
                    f"- Expected policy: {item['expected_policy'] or '-'}",
                    f"- Selected policy: {diagnostics['prompt'].get('selected_policy', '-')}",
                    f"- Issues: {', '.join(item['issues']) or '-'}",
                    f"- Supervisor execution: {supervisor.get('execution_kind', '-')}",
                    f"- Selected agents: {', '.join(supervisor.get('selected_agents') or []) or '-'}",
                    f"- Needs clarification: {supervisor.get('needs_clarification', False)}",
                    f"- Graph path: {' > '.join(supervisor.get('graph_path') or []) or '-'}",
                    "",
                    "**RAG Context**",
                    "",
                    f"- Backend: {rag_backend or '-'}",
                    f"- Hits: {len(rag_context_items)}",
                ]
            )
            if rag_context_items:
                lines.append("")
                for idx, rag_item in enumerate(rag_context_items, start=1):
                    lines.append(f"{idx}. {rag_item}")
            else:
                lines.extend(["", "- Нет RAG-фрагментов"])
            lines.extend(
                [
                    "",
                    "**Input**",
                    "",
                    item["input_text"],
                    "",
                    "**Final Response**",
                    "",
                    item["response"],
                    "",
                    "**Timeline**",
                    "",
                ]
            )
            for row in _extract_case_timeline_rows(item):
                lines.extend(
                    [
                        f"{row['step_no']}. `{row['step_label']}` / `{row['actor']}` / status=`{row['step_status']}`",
                        f"latency_ms={row['latency_ms']} tokens_in={row['tokens_input']} tokens_out={row['tokens_output']} prompt_chars={row['prompt_chars']}",
                        f"reasons: {row['reasons'] or '-'}",
                        f"warnings: {row['warnings'] or '-'}",
                        f"context: {row['selected_context_sections'] or '-'}",
                    ]
                )
                details = str(row["details_json"]).strip()
                if details and details != "{}":
                    lines.extend(["```json", details, "```"])
            lines.append("")

    if isinstance(results, dict):
        for run_name, run_results in results.items():
            append_run(run_name, run_results)
    else:
        append_run(None, results)

    return "\n".join(lines).rstrip() + "\n"


def _write_markdown_report(content: str, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(content, encoding="utf-8")


def _build_json_payload(
    results: list[dict] | dict[str, list[dict]],
    *,
    generated_at: str,
    patient_id: int,
    source: str,
    cases_path: Path,
    xlsx_path: Path,
) -> dict:
    if isinstance(results, dict):
        runs_payload: dict[str, object] = {}
        for tier_name, tier_results in results.items():
            statuses = Counter(item["status"] for item in tier_results)
            issues = Counter(issue for item in tier_results for issue in item["issues"])
            policies = Counter(item["diagnostics"].get("prompt", {}).get("selected_policy", "unknown") for item in tier_results)
            runs_payload[tier_name] = {
                "summary": {
                    "statuses": dict(statuses),
                    "issues": dict(issues),
                    "policies": dict(policies),
                    "avg_total_stage_latency_ms": round(
                        _safe_mean([int(item["diagnostics"]["summary"].get("total_stage_latency_ms", 0)) for item in tier_results]), 1
                    ),
                    "avg_llm_latency_ms": round(
                        _safe_mean([int(item["diagnostics"]["llm_call"].get("latency_ms", 0)) for item in tier_results]), 1
                    ),
                    "avg_tokens_input": round(_safe_mean([int(item["tokens_input"]) for item in tier_results]), 1),
                    "avg_tokens_output": round(_safe_mean([int(item["tokens_output"]) for item in tier_results]), 1),
                },
                "results": tier_results,
            }

        return {
            "meta": {
                "generated_at": generated_at,
                "patient_id": patient_id,
                "source": source,
                "cases_path": str(cases_path),
                "xlsx_report": str(xlsx_path),
                "cases_total": sum(len(items) for items in results.values()),
                "compare_model_tiers": list(results.keys()),
            },
            "runs": runs_payload,
        }

    statuses = Counter(item["status"] for item in results)
    issues = Counter(issue for item in results for issue in item["issues"])
    policies = Counter(item["diagnostics"].get("prompt", {}).get("selected_policy", "unknown") for item in results)

    return {
        "meta": {
            "generated_at": generated_at,
            "patient_id": patient_id,
            "source": source,
            "cases_path": str(cases_path),
            "xlsx_report": str(xlsx_path),
            "cases_total": len(results),
        },
        "summary": {
            "statuses": dict(statuses),
            "issues": dict(issues),
            "policies": dict(policies),
            "avg_total_stage_latency_ms": round(
                _safe_mean([int(item["diagnostics"]["summary"].get("total_stage_latency_ms", 0)) for item in results]), 1
            ),
            "avg_llm_latency_ms": round(
                _safe_mean([int(item["diagnostics"]["llm_call"].get("latency_ms", 0)) for item in results]), 1
            ),
            "avg_tokens_input": round(_safe_mean([int(item["tokens_input"]) for item in results]), 1),
            "avg_tokens_output": round(_safe_mean([int(item["tokens_output"]) for item in results]), 1),
        },
        "results": results,
    }


def _write_json_report(payload: dict, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(_json_dump(payload), encoding="utf-8")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run LLM chat quality evals and export an Excel report.")
    parser.add_argument("--cases", type=Path, default=DEFAULT_CASES_PATH, help="Path to YAML file with eval cases.")
    parser.add_argument("--patient-id", type=int, default=1, help="Patient ID used for real context building.")
    parser.add_argument("--source", type=str, default="text", choices=["text", "button", "system"], help="Router source.")
    parser.add_argument("--override-model-tier", type=str, default=None, choices=["lite", "pro", "max"], help="Force one model tier for all eval cases.")
    parser.add_argument("--compare-model-tiers", nargs="+", choices=["lite", "pro", "max"], help="Run the same eval set on multiple model tiers and build a comparison report.")
    parser.add_argument("--case-id", type=str, default=None, help="Run only one eval case by id.")
    parser.add_argument("--output", type=Path, default=None, help="Optional output .xlsx path.")
    return parser.parse_args()


async def _main() -> None:
    args = _parse_args()
    load_environment()
    cases = _load_cases(args.cases)
    cases = _filter_cases(cases, args.case_id)
    if not cases:
        raise RuntimeError(f"No cases found in {args.cases}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = args.output or (DEFAULT_REPORTS_DIR / f"llm_eval_{timestamp}.xlsx")
    json_output_path = output_path.with_suffix(".json")
    md_output_path = output_path.with_suffix(".md")
    if args.override_model_tier and args.compare_model_tiers:
        raise RuntimeError("Use either --override-model-tier or --compare-model-tiers, not both.")

    if args.compare_model_tiers:
        runs: dict[str, list[dict]] = {}
        for tier_name in args.compare_model_tiers:
            _run_eval_cases._override_model_tier = tier_name
            runs[tier_name] = await _run_eval_cases(cases, patient_id=args.patient_id, source=args.source)
        results_for_output: list[dict] | dict[str, list[dict]] = runs
    else:
        _run_eval_cases._override_model_tier = args.override_model_tier
        results_for_output = await _run_eval_cases(cases, patient_id=args.patient_id, source=args.source)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    json_payload = _build_json_payload(
        results_for_output,
        generated_at=generated_at,
        patient_id=args.patient_id,
        source=args.source,
        cases_path=args.cases,
        xlsx_path=output_path,
    )
    _write_json_report(json_payload, json_output_path)
    markdown_report = _build_markdown_report(results_for_output, generated_at=generated_at, patient_id=args.patient_id)
    _write_markdown_report(markdown_report, md_output_path)
    print(f"Excel report skipped: {output_path}")
    print(f"JSON report written to: {json_output_path}")
    print(f"Markdown report written to: {md_output_path}")


if __name__ == "__main__":
    asyncio.run(_main())

